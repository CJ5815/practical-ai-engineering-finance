"""A multi-year DCF/projection model: pure Python functions, plus an Excel
workbook builder that mirrors the same math as live formulas.

Philosophy-agnostic (like financial_model.py) — any future skill can reuse
this, not just the value investor. Assumes the reader already knows DCF
fundamentals; this module is about building one reproducibly, not teaching
what a DCF is.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from pydantic import BaseModel


class DCFAssumptions(BaseModel):
    """Every input the DCF needs, in one place — nothing else hardcodes a number."""

    base_revenue: float
    revenue_growth_rates: list[float]
    """One growth rate per projected year, e.g. [0.08, 0.07, 0.06, 0.05, 0.04]."""
    ebit_margin: float
    tax_rate: float
    da_pct_revenue: float
    capex_pct_revenue: float
    nwc_change_pct_revenue: float
    wacc: float
    terminal_growth_rate: float
    shares_outstanding: float
    net_debt: float
    current_share_price: float


def project_revenue(base_revenue: float, growth_rates: list[float]) -> list[float]:
    """Project revenue forward one year per growth rate.

    Args:
        base_revenue: The most recent actual year's revenue.
        growth_rates: One growth rate per projected year.

    Returns:
        Projected revenue for each year, same length as growth_rates.
    """
    revenues = []
    previous = base_revenue
    for rate in growth_rates:
        previous = previous * (1 + rate)
        revenues.append(previous)
    return revenues


def unlevered_fcf(
    revenues: list[float],
    ebit_margin: float,
    tax_rate: float,
    da_pct_revenue: float,
    capex_pct_revenue: float,
    nwc_change_pct_revenue: float,
) -> list[float]:
    """Calculate unlevered free cash flow for each projected year.

    FCF = EBIT * (1 - tax_rate) + D&A - Capex - change in net working capital,
    with D&A, Capex, and the NWC change each expressed as a % of that year's revenue.
    """
    fcf = []
    for revenue in revenues:
        ebit = revenue * ebit_margin
        nopat = ebit * (1 - tax_rate)
        da = revenue * da_pct_revenue
        capex = revenue * capex_pct_revenue
        nwc_change = revenue * nwc_change_pct_revenue
        fcf.append(nopat + da - capex - nwc_change)
    return fcf


def discount_factors(wacc: float, num_years: int) -> list[float]:
    """Calculate a discount factor for each year, 1 through num_years."""
    return [1 / (1 + wacc) ** year for year in range(1, num_years + 1)]


def terminal_value(final_year_fcf: float, wacc: float, terminal_growth_rate: float) -> float:
    """Gordon growth terminal value as of the end of the final projected year."""
    return final_year_fcf * (1 + terminal_growth_rate) / (wacc - terminal_growth_rate)


def intrinsic_value_per_share(assumptions: DCFAssumptions) -> float:
    """Run the full DCF and return intrinsic value per share.

    This is the pure-Python source of truth that build_dcf_workbook's
    Excel formulas are checked against (tests/test_dcf_model.py).
    """
    revenues = project_revenue(assumptions.base_revenue, assumptions.revenue_growth_rates)
    fcfs = unlevered_fcf(
        revenues,
        assumptions.ebit_margin,
        assumptions.tax_rate,
        assumptions.da_pct_revenue,
        assumptions.capex_pct_revenue,
        assumptions.nwc_change_pct_revenue,
    )
    factors = discount_factors(assumptions.wacc, len(fcfs))
    pv_fcfs = sum(fcf * factor for fcf, factor in zip(fcfs, factors))

    tv = terminal_value(fcfs[-1], assumptions.wacc, assumptions.terminal_growth_rate)
    pv_tv = tv * factors[-1]

    enterprise_value = pv_fcfs + pv_tv
    equity_value = enterprise_value - assumptions.net_debt
    return equity_value / assumptions.shares_outstanding


def margin_of_safety_pct(intrinsic_value: float, current_price: float) -> float:
    """How far below intrinsic value the current price sits, as a percentage.

    Positive means the stock trades below intrinsic value (a margin of
    safety, in Graham/Buffett terms); negative means it trades above.
    """
    return (intrinsic_value - current_price) / intrinsic_value


def build_dcf_workbook(assumptions: DCFAssumptions, path: str) -> None:
    """Write a multi-year DCF as a real Excel workbook with live formulas.

    Every formula uses only plain arithmetic and SUM (no NPV/IRR), so each
    year's math is visible and auditable, and the workbook can be
    recalculated by a plain formula engine for testing (test_dcf_model.py).

    An Assumptions sheet holds every input as a named cell; the DCF Model
    sheet references those cells rather than hardcoding numbers, mirroring
    the pure functions above exactly.

    Args:
        assumptions: Every input the model needs.
        path: Where to save the .xlsx file.
    """
    num_years = len(assumptions.revenue_growth_rates)
    year_columns = [get_column_letter(i + 2) for i in range(num_years)]  # B, C, D, ...
    last_col = year_columns[-1]

    workbook = Workbook()
    _write_assumptions_sheet(workbook, assumptions, year_columns)
    _write_dcf_sheet(workbook, year_columns, last_col)
    workbook.save(path)


def _write_assumptions_sheet(
    workbook: Workbook, assumptions: DCFAssumptions, year_columns: list[str]
) -> None:
    ws = workbook.active
    ws.title = "Assumptions"

    ws["A1"] = "Assumptions"
    labeled_values = [
        ("Base Revenue", assumptions.base_revenue),
        ("EBIT Margin", assumptions.ebit_margin),
        ("Tax Rate", assumptions.tax_rate),
        ("D&A % of Revenue", assumptions.da_pct_revenue),
        ("Capex % of Revenue", assumptions.capex_pct_revenue),
        ("NWC Change % of Revenue", assumptions.nwc_change_pct_revenue),
        ("WACC", assumptions.wacc),
        ("Terminal Growth Rate", assumptions.terminal_growth_rate),
        ("Shares Outstanding", assumptions.shares_outstanding),
        ("Net Debt", assumptions.net_debt),
        ("Current Share Price", assumptions.current_share_price),
    ]
    for row, (label, value) in enumerate(labeled_values, start=2):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value

    ws["A14"] = "Revenue Growth Rates by Year"
    for col, rate in zip(year_columns, assumptions.revenue_growth_rates):
        ws[f"{col}15"] = f"Year {year_columns.index(col) + 1}"
        ws[f"{col}16"] = rate

    ws.column_dimensions["A"].width = 26


# Assumptions sheet row numbers, named for readability in the formulas below.
_BASE_REVENUE_ROW = 2
_EBIT_MARGIN_ROW = 3
_TAX_RATE_ROW = 4
_DA_PCT_ROW = 5
_CAPEX_PCT_ROW = 6
_NWC_PCT_ROW = 7
_WACC_ROW = 8
_TERMINAL_GROWTH_ROW = 9
_SHARES_ROW = 10
_NET_DEBT_ROW = 11
_PRICE_ROW = 12
_GROWTH_RATES_ROW = 16


def _write_dcf_sheet(workbook: Workbook, year_columns: list[str], last_col: str) -> None:
    ws = workbook.create_sheet("DCF Model")

    ws["A1"] = "Year"
    for i, col in enumerate(year_columns, start=1):
        ws[f"{col}1"] = i
        ws[f"{col}1"].number_format = '"Year "0'

    ws["A2"] = "Revenue"
    ws["A3"] = "EBIT"
    ws["A4"] = "NOPAT"
    ws["A5"] = "D&A"
    ws["A6"] = "Capex"
    ws["A7"] = "NWC Change"
    ws["A8"] = "Unlevered FCF"
    ws["A9"] = "Discount Factor"
    ws["A10"] = "PV of FCF"

    for i, col in enumerate(year_columns):
        prior_revenue = f"Assumptions!$B${_BASE_REVENUE_ROW}" if i == 0 else f"{year_columns[i - 1]}2"
        growth_cell = f"Assumptions!{col}{_GROWTH_RATES_ROW}"
        ws[f"{col}2"] = f"={prior_revenue}*(1+{growth_cell})"
        ws[f"{col}3"] = f"={col}2*Assumptions!$B${_EBIT_MARGIN_ROW}"
        ws[f"{col}4"] = f"={col}3*(1-Assumptions!$B${_TAX_RATE_ROW})"
        ws[f"{col}5"] = f"={col}2*Assumptions!$B${_DA_PCT_ROW}"
        ws[f"{col}6"] = f"={col}2*Assumptions!$B${_CAPEX_PCT_ROW}"
        ws[f"{col}7"] = f"={col}2*Assumptions!$B${_NWC_PCT_ROW}"
        ws[f"{col}8"] = f"={col}4+{col}5-{col}6-{col}7"
        ws[f"{col}9"] = f"=1/(1+Assumptions!$B${_WACC_ROW})^{col}1"
        ws[f"{col}10"] = f"={col}8*{col}9"

    ws["A12"] = "Sum PV of FCF"
    ws["B12"] = f"=SUM(B10:{last_col}10)"

    ws["A13"] = "Terminal Value"
    ws["B13"] = (
        f"={last_col}8*(1+Assumptions!$B${_TERMINAL_GROWTH_ROW})"
        f"/(Assumptions!$B${_WACC_ROW}-Assumptions!$B${_TERMINAL_GROWTH_ROW})"
    )

    ws["A14"] = "PV of Terminal Value"
    ws["B14"] = f"=B13*{last_col}9"

    ws["A15"] = "Enterprise Value"
    ws["B15"] = "=B12+B14"

    ws["A16"] = "Less: Net Debt"
    ws["B16"] = f"=-Assumptions!$B${_NET_DEBT_ROW}"

    ws["A17"] = "Equity Value"
    ws["B17"] = "=B15+B16"

    ws["A18"] = "Shares Outstanding"
    ws["B18"] = f"=Assumptions!$B${_SHARES_ROW}"

    ws["A19"] = "Intrinsic Value Per Share"
    ws["B19"] = "=B17/B18"

    ws["A20"] = "Current Share Price"
    ws["B20"] = f"=Assumptions!$B${_PRICE_ROW}"

    ws["A21"] = "Margin of Safety %"
    ws["B21"] = "=(B19-B20)/B19"

    ws.column_dimensions["A"].width = 22
